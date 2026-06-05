"""Inspect mehrfachversicherte-RA workbooks: sheet names and first 15 rows of each sheet."""
import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIR = os.path.join(ROOT, 'data', 'statistik-mehrfachversicherte-ra')

files = sorted(os.listdir(DIR))
for fname in files:
    path = os.path.join(DIR, fname)
    print(f"\n{'='*70}")
    print(f"FILE: {fname}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True, max_row=20))
        print(f"\n  Sheet: {sheet_name!r}")
        for i, row in enumerate(rows):
            non_null = [c for c in row if c is not None]
            if non_null:
                print(f"    row{i:02d}: {row}")
    wb.close()
